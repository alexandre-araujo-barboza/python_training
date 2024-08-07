# openpyxl - criando uma planilha do Excel (Workbook e Worksheet)
# openpyxl - manipulando as planilhas do Workbook

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active

# Nome para a planilha
sheet_name = 'Planilha de exemplo'

# Criamos a planilha
workbook.create_sheet(sheet_name, 0)
print('Workbook created.')

# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

# Remover uma planilha
workbook.remove(workbook['Sheet'])
print('Workbook removed.')

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
print(f'Cell {worksheet.active_cell} was created.'
      f' on {workbook.active}') 

